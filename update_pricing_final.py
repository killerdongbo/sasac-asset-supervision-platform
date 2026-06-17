"""
Final: 精确逐行写，用硬编码行号避开所有编码解析问题。
"""
import shutil, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

SRC = "功能清单与报价/湛江市国资监管平台升级报价-20260616.xlsx"
DST = "功能清单与报价/湛江市国资监管平台升级报价-20260616-更新版.xlsx"
shutil.copy(SRC, DST)

wb = openpyxl.load_workbook(DST)
ws = wb[wb.sheetnames[0]]

# 找到投资管理所在行
invest_row = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 2).value
    if v and '投资管理' in str(v):  # 投资管理
        invest_row = r
        break
print(f"invest_row={invest_row}")

# 插入7行
for _ in range(7):
    ws.insert_rows(invest_row + 1)

# 写入数据 — 使用显式行号
data = [
    # row, col, value
    # 标题行
    (invest_row+1, 1, "二期升级模块（Phase 2 — 监管端深化升级）"),
    (invest_row+1, 6, "原报价已覆盖"),
    # 模块1: 产权管理
    (invest_row+2, 2, "产权管理系统"),
    (invest_row+2, 3, "产权占有/变动/注销登记，电子产权证自动生成（含二维码验证），产权树状图可视化，产权交易实时监测（对接全国产权交易机构），资产评估备案管理（自动计算评估偏差率），产权档案电子归档"),
    (invest_row+2, 4, "企业国有产权全生命周期数字化管理。对接全国产权交易机构实时抓取挂牌/竞价/成交信息，交易价格偏离评估值>15%自动预警。产权层级树状图穿透展示国资委-集团-子公司-孙公司控制链条。"),
    (invest_row+2, 5, "已含"),
    (invest_row+2, 6, "监管升级模块"),
    # 模块2: 财务监督
    (invest_row+3, 2, "财务监督系统"),
    (invest_row+3, 3, "财务数据自动采集（对接企业ERP/财务系统），三大财务报表在线填报与汇总，预算执行动态监控，大额资金实时监控（单笔>100万或日累计>500万自动标记），关键财务指标自动计算（资产负债率/ROE/周转率等），财务风险分级预警，合并报表自动生成"),
    (invest_row+3, 4, "通过API对接企业ERP/财务系统自动采集资产负债表/利润表/现金流量表数据。大额资金支付系统实时校验风控规则。资产负债率>80%红色预警、连续两年亏损预警、经营现金流恶化预警等多维度财务健康监测。"),
    (invest_row+3, 5, "已含"),
    (invest_row+3, 6, "监管升级模块"),
    # 模块3: 业绩考核
    (invest_row+4, 2, "业绩考核与薪酬管理"),
    (invest_row+4, 3, "考核指标体系配置（经济指标/管理指标/分类指标），指标下达与动态监测，年度/任期考核自动评分（目标vs实际加权计算），企业负责人薪酬总额与考核结果挂钩调整，中长期激励方案（股权/分红）管理"),
    (invest_row+4, 4, "经营业绩考核指标下达-动态监测-自动评分-薪酬挂钩全闭环管理。考核未达标系统自动调减薪酬总额。支持KPI/OKR/360度多种考核方案，权重合计100%校验。"),
    (invest_row+4, 5, "已含"),
    (invest_row+4, 6, "监管升级模块"),
    # 模块4: 监督追责
    (invest_row+5, 2, "监督追责系统"),
    (invest_row+5, 3, "内部审计全流程（计划/作业/报告/整改），违规经营投资线索受理-初步核实-立案调查-责任认定-处理决定全流程，国有资产损失量化认定，审计整改超期自动升级预警，纪检监察/巡视巡察/外部审计信息协同，违规案例库建设与智能预警"),
    (invest_row+5, 4, "审计计划-实施-报告-整改-销号闭环管理。违规追责全程留痕不可篡改（独立审计库存储）。追责结果自动关联薪酬扣减和职务任免建议。案例特征自动匹配新线索辅助智能预警。"),
    (invest_row+5, 5, "已含"),
    (invest_row+5, 6, "监管升级模块"),
    # 模块5: 重大事项
    (invest_row+6, 2, "重大事项管理"),
    (invest_row+6, 3, "企业合并/分立/改制/上市/破产等重大事项报告与审批，事项处理进度动态跟踪，重大诉讼案件全流程管理（立案/审理/判决/执行），对外担保审批与风险监控（到期30天内自动预警），全过程材料电子归档"),
    (invest_row+6, 4, "与三重一大决策管理模块互补：三重一大侧重决策过程合规监督，重大事项管理侧重事项全过程跟踪与风险防范。担保风险分级（高/中/低）动态管理。"),
    (invest_row+6, 5, "已含"),
    (invest_row+6, 6, "监管升级模块"),
    # 模块6: 预警驾驶舱 + 移动端
    (invest_row+7, 2, "预警驾驶舱升级 + 移动端重做"),
    (invest_row+7, 3, "综合预警体系升级（覆盖资产/人力/项目/投资/财务/合规/监督7大维度），决策驾驶舱升级（5大维度关键指标一屏总览+预警红灯区+跨模块下钻穿透），移动端重构为独立uni-app项目（8大功能模块：HR审批/三重一大/项目进度填报/投资查询/财务快报/预警消息/制度文件/个人中心），支持微信小程序/H5/APP多端发布"),
    (invest_row+7, 4, "预警引擎从一期资产单一维度扩展至全业务7大维度。驾驶舱支持从宏观汇总指标点击下钻穿透至底层明细数据。移动端从H5内嵌升级为独立uni-app项目，新增6类业务场景移动化，支持微信小程序/H5/APP多端发布。"),
    (invest_row+7, 5, "已含"),
    (invest_row+7, 6, "升级优化模块"),
]

blue = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
for r, c, v in data:
    cell = ws.cell(row=r, column=c)
    cell.value = v
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    is_title = (c == 1 and "二期升级" in str(v))
    cell.font = Font(name='微软雅黑', size=11 if is_title else 10, bold=is_title,
                     color='1F4E79' if is_title else '000000')
    if is_title:
        cell.fill = blue

# 更新建设周期
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v and '建设周期' in str(v):
        ws.cell(r, 1).value = "      3. 建设周期：一期（固定资产核心模块）4周，二期（4大核心业务模块+6大升级模块+移动端重做）22周，合计26周。"

# 列宽
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 48
ws.column_dimensions['D'].width = 52
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 22

wb.save(DST)
print(f"Saved: {DST}")

# Final verification
wb2 = openpyxl.load_workbook(DST, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
checks = {
    '产权管理': False,  # 产权管理
    '财务监督': False,  # 财务监督
    '业绩考核': False,  # 业绩考核
    '监督追责': False,  # 监督追责
    '重大事项': False,  # 重大事项
    '预警驾驶舱': False,  # 预警驾驶舱
    '硬件': False,  # 硬件
    '信创': False,  # 信创
}
for r in range(1, ws2.max_row + 1):
    row_text = ''.join([str(ws2.cell(r, c).value or '') for c in range(1, 7)])
    for k in checks:
        if k in row_text:
            checks[k] = True

all_ok = True
for k, v in checks.items():
    status = "OK" if v else "MISSING!"
    if not v: all_ok = False
    print(f"  {k}: {status}")

if all_ok:
    print("\nALL CHECKS PASSED!")
else:
    print("\nSOME CHECKS FAILED!")

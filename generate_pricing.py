"""
Generate pricing Excel for 湛江市国资监管平台升级方案.
Sheet 1: Platform quote (~300万) with existing + new features
Sheet 2: Value-added services (~100万) from PPT slide 33
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from copy import copy

DST = r"d:\谢东波个人\智算\国资监管平台\湛江市国资监管平台升级报价-20260616.xlsx"

wb = openpyxl.Workbook()

# ── Style definitions ──
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
title_font = Font(name='微软雅黑', size=14, bold=True, color='1A56D6')
subtitle_font = Font(name='微软雅黑', size=10, color='666666')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=10, bold=True)
total_font = Font(name='微软雅黑', size=12, bold=True, color='C00000')
category_font = Font(name='微软雅黑', size=11, bold=True, color='1A56D6')
header_fill = PatternFill(start_color='3C89FF', end_color='3C89FF', fill_type='solid')
category_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
light_gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, max_col, fill=None, font=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if fill: cell.fill = fill
        if font: cell.font = font
        if col in (5, 6): cell.alignment = center_align
        elif col == 7: cell.alignment = center_align
        else: cell.alignment = left_align

# ================================================================
# SHEET 1: 平台报价
# ================================================================
ws1 = wb.active
ws1.title = '平台建设报价'

# Column widths
col_widths = [16, 28, 42, 52, 10, 14, 18]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title row
ws1.merge_cells('A1:G1')
c = ws1.cell(row=1, column=1, value='湛江市国资监管平台升级方案报价单')
c.font = title_font; c.alignment = center_align

ws1.merge_cells('A2:G2')
c = ws1.cell(row=2, column=1, value='涵盖固定资产全生命周期管理 + 人事/会议/项目/投资四大新增模块 | 信创适配 · 二级等保 · SaaS多租户')
c.font = subtitle_font; c.alignment = center_align

ws1.row_dimensions[1].height = 32
ws1.row_dimensions[2].height = 22

# Headers (row 4)
headers = ['费用项目', '业务套件名称', '包含功能', '功能说明', '配置', '费用（万元）', '备注']
for col, h in enumerate(headers, 1):
    ws1.cell(row=4, column=col, value=h)
style_header(ws1, 4, 7)

# ── Pricing data ──
platform_items = [
    # (category, suite_name, features, description, config, price, remark, is_category)
    ('固定资产全生命周期管理平台', None, None, None, None, None, None, 'category'),
    (None, '固定资产核心管理套件',
     '系统入口与安全认证、组织权限管理（RBAC四角色）、数据批量导入导出、资产基础资料（分类/位置/供应商）、资产台账管理（登记/编号/标签/变更记录）、资产流转管理（入库/领用/调拨/报废）、资产标签与二维码打印',
     '覆盖资产从入账到报废的全生命周期管理，支持一物一码精准定位、权属灵活重划、全链追溯。结合多租户SaaS架构，实现市属五大集团及下属企业统一平台、数据隔离。',
     '1套', 95, '核心基础套件（含18个功能模块）'),
    (None, '采购与财务一体化套件',
     '采购管理（申请/验收/转资产）、财务折旧管理（规则配置/计提/确认）、财务系统对接适配器',
     '打通采购→验收→资产入账全流程自动化。内置直线法/双倍余额递减法等多种折旧方案，支持与用友/金蝶等财务系统数据对接，实现业财一体化。',
     '1套', 30, '业财协同套件'),
    (None, '巡检盘点与维保管理套件',
     '巡检管理（任务制定/现场登记/异常识别）、盘点管理（任务/扫码登记/差异自动识别/报告）、维保维修管理（申请/派单/维修登记/巡检异常自动触发）',
     '建立\"巡检→异常→维修\"闭环管理，支持移动端现场扫码登记。盘点差异自动计算（账面vs实盘），生成盘点报告。巡检异常自动触发维修工单，减少资产流失。',
     '1套', 28, '资产运维套件'),
    (None, '国资监管与智能预警套件',
     '管理驾驶舱与工作台、统计报表（资产/巡检/盘点/折旧）、监管上报（19类报表采集/审核/版本快照）、审批工作流引擎（可配置/条件分支）、预警中心（维保到期/巡检超期/闲置资产）、审计追溯（操作日志/全生命周期追踪）',
     '面向国资委监管需求，提供管理驾驶舱一屏掌控全局、19类国资报表自动采集上报、可配置审批工作流引擎。内置多维度智能预警规则，主动推送通知。全链路审计日志基于JSONB快照独立存储、不可篡改。',
     '1套', 32, '监管决策套件'),
    ('新增业务模块', None, None, None, None, None, None, 'category'),
    (None, '人力资源管理套件（新增）',
     '组织架构管理、员工档案管理（入职/转正/调岗/离职）、薪酬福利管理、考勤管理、绩效考核、招聘管理、培训管理、劳动合同管理、人力成本分析报表',
     '整合国企\"选、育、用、留\"全生命周期管理，建立标准化人事流程。支持多级组织架构、岗位编制管控、薪酬自动核算、绩效考核闭环。自动生成人力成本分析报表，辅助科学决策。',
     '1套', 28, '新增核心模块'),
    (None, '三重一大会议管理套件（新增）',
     '议题申报与审核、会议通知与签到、会议纪要生成与分发、决议事项跟踪督办、决策过程全程留痕、多端协同（PC/移动/大屏）、统计分析报表',
     '落实国企\"三重一大\"决策制度要求，实现重大决策事项的全流程数字化管理。覆盖议题申报→审核→上会→决议→督办→归档全闭环，决策过程全程留痕、可追溯、可审计，满足巡视审计要求。',
     '1套', 22, '新增核心模块'),
    (None, '项目管理套件（新增）',
     '项目立项与可研管理、项目计划与进度跟踪、项目预算与成本管控、项目验收与后评价、项目合同管理、项目文档管理、项目看板与甘特图、专项债项目管理',
     '实现企业常规项目与国企专项债项目的全生命周期管理，覆盖立项→执行→监控→验收→后评价全流程。支持甘特图/看板可视化、预算执行预警、合同履约跟踪、项目成果归档。',
     '1套', 25, '新增核心模块'),
    (None, '投资管理套件（新增）',
     '投资计划编制与审批、投资项目库管理、投前尽职调查与风险评估、投中交易管理与资金监管、投后跟踪管理与退出管理、投资回报分析、股权投资穿透式管理',
     '建立国企投资全生命周期管控体系，覆盖投前→投中→投后全流程。支持股权投资穿透式监管（层层追溯至底层资产）、投资回报率动态测算、风险预警分级管理，满足国资委对国企投资合规监管要求。',
     '1套', 23, '新增核心模块'),
    ('基础设施与技术服务', None, None, None, None, None, None, 'category'),
    (None, '硬件服务器与网络设备',
     '应用服务器×2（国产信创服务器，鲲鹏/飞腾CPU、32核/128G/2TB SSD）、数据库服务器×1（32核/256G/4TB SSD RAID5）、万兆交换机×2（堆叠部署）、下一代防火墙×1（含IPS/AV模块）、UPS不间断电源（6KVA/4H）',
     '提供国产化信创硬件底座，满足二级等保对物理安全、网络安全、主机安全的要求。应用服务器双机热备，数据库服务器RAID5冗余，确保系统7×24小时稳定运行。网络设备万兆互联，防火墙提供入侵防御与病毒防护。',
     '1套', 20, '硬件基础设施（一次性）'),
    (None, '信创适配与云基础设施',
     '国产操作系统适配（麒麟/统信）、国产数据库适配（达梦DM8/人大金仓）、国产中间件适配（东方通/中创）、ARM/x86混合架构部署、云服务器资源（1年）、二级等保合规建设',
     '全面适配信创生态环境，确保系统在国产化环境下性能零损耗。提供云基础设施底座与二级等保合规建设，满足监管部门对云上国资系统的合规要求。',
     '1年', 18, '基础设施（按年）'),
    (None, '定制化开发与系统集成',
     '个性化功能定制、业务流程适配、第三方系统集成对接（ERP/财务/OA）、持续迭代优化服务',
     '基于客户个性化业务需求，提供灵活的定制化功能开发与系统集成服务，按需匹配开发资源，确保系统与现有IT生态无缝衔接。',
     '按需', '15-25', '1000元/人天'),
]

row = 5
for item in platform_items:
    if len(item) > 7 and item[7] == 'category':
        # Category header row
        ws1.merge_cells(f'A{row}:G{row}')
        c = ws1.cell(row=row, column=1, value=item[0])
        c.font = category_font; c.fill = category_fill; c.alignment = left_align
        ws1.row_dimensions[row].height = 24
        for col in range(1, 8):
            ws1.cell(row=row, column=col).border = thin_border
            ws1.cell(row=row, column=col).fill = category_fill
        row += 1
    else:
        ws1.cell(row=row, column=1).value = item[0] or ''
        ws1.cell(row=row, column=2).value = item[1]
        ws1.cell(row=row, column=3).value = item[2]
        ws1.cell(row=row, column=4).value = item[3]
        ws1.cell(row=row, column=5).value = item[4]
        ws1.cell(row=row, column=6).value = item[5]
        ws1.cell(row=row, column=7).value = item[6]
        ws1.row_dimensions[row].height = 80
        style_row(ws1, row, 7)
        row += 1

# Total row
ws1.merge_cells(f'A{row}:E{row}')
c = ws1.cell(row=row, column=1, value='合  计')
c.font = total_font; c.alignment = center_align
c.fill = total_fill
c = ws1.cell(row=row, column=6, value='316 - 326')
c.font = total_font; c.alignment = center_align; c.fill = total_fill
c = ws1.cell(row=row, column=7, value='≈ 320万元（平台建设）')
c.font = total_font; c.alignment = center_align; c.fill = total_fill
style_row(ws1, row, 7, total_fill, total_font)
ws1.row_dimensions[row].height = 30
row += 1

# Notes
row += 1
notes = [
    '注：1. 以上报价不含云服务器、短信、邮箱、WPS对接等第三方服务费用（按甲方需要另议）。',
    '      2. 报价有效期为30天，最终价格以双方确认的功能范围和实施计划为准。',
    '      3. 交付周期：一期（固定资产基础版）4周，二期（全模块+新增功能）10周，合计14周。',
    '      4. 质保期1年，质保期内免费修复缺陷；后续年度运维费按合同金额10%/年计收。',
]
for note in notes:
    ws1.merge_cells(f'A{row}:G{row}')
    c = ws1.cell(row=row, column=1, value=note)
    c.font = Font(name='微软雅黑', size=9, color='999999')
    c.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

# ================================================================
# SHEET 2: 增值服务
# ================================================================
ws2 = wb.create_sheet('增值运营服务')

col_widths2 = [16, 26, 48, 12, 10, 16]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Title
ws2.merge_cells('A1:F1')
c = ws2.cell(row=1, column=1, value='湛江市国资监管平台 —— 增值运营服务报价')
c.font = title_font; c.alignment = center_align

ws2.merge_cells('A2:F2')
c = ws2.cell(row=2, column=1, value='涵盖数据治理、平台运营、人才陪跑、资产盘点、经营诊断、制度优化六大增值服务 | 三年全程陪伴式服务')
c.font = subtitle_font; c.alignment = center_align

ws2.row_dimensions[1].height = 32
ws2.row_dimensions[2].height = 22

# Headers
headers2 = ['服务类别', '服务项名称', '具体服务内容', '服务周期', '费用（万元）', '备注']
for col, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=col, value=h)
style_header(ws2, 4, 6)

# ── Value-added services data ──
value_added_items = [
    ('数据治理', None, None, None, None, None, 'category'),
    (None, '资产数据初始化与清洗',
     '对五大集团及下属企业的存量资产数据进行全面清洗、去重、校验与标准化处理，确保迁移至新系统的数据零差错。包括：历史数据格式转换、异常数据识别与修正、资产分类标准统一、资产编码体系重建。',
     '一次性', 18, '上线前必须完成'),
    (None, '数据质量持续治理',
     '建立数据质量监控体系，定期扫描数据完整性/一致性/准确性指标。每季度出具数据质量报告，发现问题自动生成修正工单。三年内持续跟踪优化，确保数据准确率≥99%。',
     '3年', 12, '4万/年×3年'),
    ('平台运营支撑', None, None, None, None, None, 'category'),
    (None, '系统运营与技术支持',
     '提供7×12小时系统运维服务，包括：日常系统监控与健康检查、性能优化与调优、安全漏洞修复、版本升级与补丁管理、数据备份与容灾恢复。配备专属运营经理，季度服务质量报告。',
     '3年', 20, '运维保障'),
    (None, '用户服务与问题响应',
     '建立ITSM工单系统，提供多渠道（电话/微信/工单）技术支持。承诺：紧急故障30分钟响应、2小时到场；一般问题2小时响应、24小时内解决。每月汇总分析高频问题，主动优化。',
     '3年', 10, 'SLA保障'),
    ('人才培养服务', None, None, None, None, None, 'category'),
    (None, '1对1人才陪跑计划',
     '项目组工程师与企业资产管理员\"一对一\"结对，采用\"我做你看→你做我看→独立操作\"三步陪跑法。确保每家市属国企至少1-2名人员能够独立完成系统运维、数据处理、报表生成等核心操作。',
     '3年', 15, '5万/年×3年'),
    (None, '分层培训体系',
     '管理层培训（系统价值、数据决策、监管要点）每年2次；操作层培训（功能操作、数据处理、常见问题）每季度1次；财务层培训（折旧、上报、对账）每半年1次。合计培训≥30场次。',
     '3年', 8, '含培训教材与考核'),
    ('资产专业服务', None, None, None, None, None, 'category'),
    (None, '资产盘点服务',
     '专业人员上门实地盘点，对土地、房屋、构筑物、设备、车辆、存货等固定资产进行全面清查核对。采用\"贴码+扫码+拍照\"三步法，出具正式盘点报告并录入系统。解决企业\"自己盘不清、盘不准\"的难题。',
     '一次性', 12, '全覆盖实地盘点'),
    (None, '资产经营诊断',
     '基于系统沉淀数据，从资产闲置率、出租回报率、资产结构合理性、价值变动趋势、折旧计提合理性等维度，对五大集团资产进行深度分析诊断。出具《资产经营诊断报告》，提出盘活建议与优化方案。',
     '一次性', 8, '数据驱动决策'),
    ('制度与可持续运营', None, None, None, None, None, 'category'),
    (None, '制度流程优化',
     '协助企业设计/优化资产管理制度、审批流程、内控规范。包括：资产管理办法、资产分类标准、折旧政策、盘点制度、处置流程、审批权限矩阵等。建立常态化资产管理机制，确保制度可落地、可持续。',
     '一次性', 6, '制度体系建设'),
    (None, '数据资产运营规划',
     '响应\"数据二十条\"政策，协助客户规划数据资产入表路径。包括：数据资产目录建立、数据确权与三权分置设计、数据资产价值评估模型、数据要素流通方案。为后续数据资产入表与价值释放奠定基础。',
     '一次性', 5, '前瞻性布局'),
]

row2 = 5
for item in value_added_items:
    if len(item) > 6 and item[6] == 'category':
        ws2.merge_cells(f'A{row2}:F{row2}')
        c = ws2.cell(row=row2, column=1, value=item[0])
        c.font = category_font; c.fill = category_fill; c.alignment = left_align
        ws2.row_dimensions[row2].height = 24
        for col in range(1, 7):
            ws2.cell(row=row2, column=col).border = thin_border
            ws2.cell(row=row2, column=col).fill = category_fill
        row2 += 1
    else:
        ws2.cell(row=row2, column=1).value = item[0] or ''
        ws2.cell(row=row2, column=2).value = item[1]
        ws2.cell(row=row2, column=3).value = item[2]
        ws2.cell(row=row2, column=4).value = item[3]
        ws2.cell(row=row2, column=5).value = item[4]
        ws2.cell(row=row2, column=6).value = item[5]
        ws2.row_dimensions[row2].height = 85
        style_row(ws2, row2, 6)
        # Alternate row shading
        if row2 % 2 == 0:
            for col in range(1, 7):
                ws2.cell(row=row2, column=col).fill = light_gray_fill
        row2 += 1

# Total row
ws2.merge_cells(f'A{row2}:D{row2}')
c = ws2.cell(row=row2, column=1, value='合  计')
c.font = total_font; c.alignment = center_align; c.fill = total_fill
c = ws2.cell(row=row2, column=5, value=114)
c.font = total_font; c.alignment = center_align; c.fill = total_fill
c = ws2.cell(row=row2, column=6, value='≈ 114万元（增值服务）')
c.font = total_font; c.alignment = center_align; c.fill = total_fill
style_row(ws2, row2, 6, total_fill, total_font)
ws2.row_dimensions[row2].height = 30
row2 += 2

# Grand total
ws2.merge_cells(f'A{row2}:D{row2}')
c = ws2.cell(row=row2, column=1, value='平台建设 + 增值服务 总合计')
c.font = Font(name='微软雅黑', size=13, bold=True, color='C00000'); c.alignment = center_align
c.fill = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
c = ws2.cell(row=row2, column=5, value='430 - 440')
c.font = Font(name='微软雅黑', size=13, bold=True, color='C00000'); c.alignment = center_align
c.fill = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
c = ws2.cell(row=row2, column=6, value='≈ 434万元（整体方案）')
c.font = Font(name='微软雅黑', size=13, bold=True, color='C00000'); c.alignment = center_align
c.fill = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
for col in range(1, 7):
    ws2.cell(row=row2, column=col).border = thin_border
    ws2.cell(row=row2, column=col).fill = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
ws2.row_dimensions[row2].height = 35
row2 += 1

# Notes
row2 += 1
notes2 = [
    '注：1. 增值服务报价可与平台建设合同一并签署，也可分阶段签署。',
    '      2. 资产盘点服务按实际资产数量计费，上表为预估（五大集团约2-3万项资产），超出部分按4元/项另计。',
    '      3. 人才陪跑与培训服务可按年度续签，续签价格维持不变。',
    '      4. 本方案未含云服务器、二级等保测评、短信/邮件等第三方服务费用，按甲方需要另议。',
    '      5. 增值服务收入部分可用于覆盖系统长期运维成本，降低财政持续投入压力。',
]
for note in notes2:
    ws2.merge_cells(f'A{row2}:F{row2}')
    c = ws2.cell(row=row2, column=1, value=note)
    c.font = Font(name='微软雅黑', size=9, color='999999')
    c.alignment = Alignment(horizontal='left', vertical='center')
    row2 += 1

# ── Print settings ──
for ws in [ws1, ws2]:
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

# Save
wb.save(DST)
print(f"DONE! Saved to: {DST}")
print(f"Sheet 1: 平台建设报价 (~300万)")
print(f"Sheet 2: 增值运营服务 (~114万)")
print(f"Total: ~410-420万")
